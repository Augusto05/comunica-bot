import discord
from discord import app_commands
from discord.ext import commands
import openpyxl
from openpyxl import Workbook
import os
import json
import glob
import asyncio
from typing import Optional

from dotenv import load_dotenv
load_dotenv()

TOKEN = os.getenv("DISCORD_TOKEN")
CANAL_ID = int(os.getenv("CANAL_ID"))
CANAL_COMANDOS_ID = int(os.getenv("CANAL_COMANDOS_ID"))


# Variáveis globais para controle do último comunicado
titulo_ultimo_comunicado = None
arquivo_ultimo_relatorio = None
json_relatorio_path = os.path.join("relatorios", "relatorio_atual.json")

# Variável global para armazenar o ID da última mensagem de comunicado
ultima_msg_comunicado_id = None

# Garante que a pasta relatorios existe
os.makedirs("relatorios", exist_ok=True)

def nome_arquivo_json(titulo):
    nome = f"{titulo}.json".replace("/", "_").replace("\\", "_").replace(":", "_").replace("*", "_").replace("?", "_").replace("\"", "_").replace("<", "_").replace(">", "_").replace("|", "_")
    return os.path.join("relatorios", nome)

def sanitizar_titulo(titulo):
    return titulo.replace("/", "_").replace("\\", "_").replace(":", "_").replace("*", "_").replace("?", "_").replace("\"", "_").replace("<", "_").replace(">", "_").replace("|", "_")

class ConfirmarLeituraView(discord.ui.View):
    def __init__(self, titulo):
        super().__init__(timeout=None)
        self.titulo = titulo

    @discord.ui.button(label="Confirmar Leitura", style=discord.ButtonStyle.success)
    async def confirmar(self, interaction: discord.Interaction, button: discord.ui.Button):
        global arquivo_ultimo_relatorio
        if not arquivo_ultimo_relatorio:
            await interaction.response.send_message("❌ Relatório não encontrado.", ephemeral=True)
            return
        wb = openpyxl.load_workbook(arquivo_ultimo_relatorio)
        ws = wb.active
        usuario_id = interaction.user.id
        for row in ws.iter_rows(min_row=2):
            if row[0].value == interaction.user.display_name:
                row[1].value = "Viu"
                break
        wb.save(arquivo_ultimo_relatorio)
        atualizar_status_json(interaction.user)
        await interaction.response.edit_message(content="✅ Leitura confirmada e registrada!", view=None)

class ComunicadoView(discord.ui.View):
    def __init__(self, titulo, conteudo):
        super().__init__(timeout=None)
        self.titulo = titulo
        self.conteudo = conteudo

    @discord.ui.button(label="Abrir Comunicado", style=discord.ButtonStyle.primary)
    async def abrir_modal(self, interaction: discord.Interaction, button: discord.ui.Button):
        embed = discord.Embed(
            title=f"📢 {self.titulo}",
            description=self.conteudo,
            color=discord.Color.blue()
        )
        # Verifica se o usuário já confirmou a leitura
        ja_leu = False
        if os.path.exists(json_relatorio_path):
            with open(json_relatorio_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            for membro in data.get("membros", []):
                if membro.get("id") == interaction.user.id and membro["status"] == "Viu":
                    ja_leu = True
                    break
        if ja_leu:
            await interaction.response.send_message(embed=embed, ephemeral=True)
        else:
            await interaction.response.send_message(embed=embed, ephemeral=True)
            # Após 25 segundos, enviar o botão de confirmação
            await asyncio.sleep(25)
            await interaction.followup.send(
                content="Se você leu o comunicado acima, confirme sua leitura:",
                view=ConfirmarLeituraView(self.titulo),
                ephemeral=True
            )

class ComunicadoViewPersistente(discord.ui.View):
    def __init__(self, titulo):
        super().__init__(timeout=None)
        self.titulo = titulo
        self.add_item(self.AbrirComunicadoButton(titulo))

    class AbrirComunicadoButton(discord.ui.Button):
        def __init__(self, titulo):
            super().__init__(
                label="Abrir Comunicado",
                style=discord.ButtonStyle.primary,
                custom_id=f"abrir_comunicado_{sanitizar_titulo(titulo)}"
            )
            self.titulo = titulo

        async def callback(self, interaction: discord.Interaction):
            json_path = nome_arquivo_json(self.titulo)
            conteudo = "[Conteúdo não encontrado]"
            if os.path.exists(json_path):
                with open(json_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    conteudo = data.get("conteudo", "[Conteúdo não encontrado]")
            embed = discord.Embed(
                title=f"📢 {self.titulo}",
                description=conteudo,
                color=discord.Color.blue()
            )
            ja_leu = False
            if os.path.exists(json_path):
                with open(json_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                for membro in data.get("membros", []):
                    if membro.get("id") == interaction.user.id and membro["status"] == "Viu":
                        ja_leu = True
                        break
            if ja_leu:
                await interaction.response.send_message(embed=embed, ephemeral=True)
            else:
                await interaction.response.send_message(embed=embed, ephemeral=True)
                await asyncio.sleep(20)
                await interaction.followup.send(
                    content="Se você leu o comunicado acima, confirme sua leitura:",
                    view=ConfirmarLeituraView(self.titulo),
                    ephemeral=True
                )

def salvar_relatorio_json(titulo, membros_status, conteudo=None):
    with open(json_relatorio_path, "w", encoding="utf-8") as f:
        json.dump({
            "titulo": titulo,
            "membros": membros_status,
            "conteudo": conteudo
        }, f, ensure_ascii=False, indent=2)
    with open(nome_arquivo_json(titulo), "w", encoding="utf-8") as f:
        json.dump({
            "titulo": titulo,
            "membros": membros_status,
            "conteudo": conteudo
        }, f, ensure_ascii=False, indent=2)

globals()["salvar_relatorio_json"] = salvar_relatorio_json

def atualizar_status_json(usuario):
    if not os.path.exists(json_relatorio_path):
        return
    with open(json_relatorio_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    usuario_id = usuario.id if hasattr(usuario, 'id') else usuario  # suporta tanto objeto quanto id direto
    for membro in data["membros"]:
        if membro.get("id") == usuario_id:
            membro["status"] = "Viu"
            break
    with open(json_relatorio_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    titulo = data["titulo"]
    with open(nome_arquivo_json(titulo), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

intents = discord.Intents.default()
intents.members = True
bot = commands.Bot(command_prefix="!", intents=intents)

# No on_ready, registrar views persistentes para todos os comunicados antigos
@bot.event
async def on_ready():
    global arquivo_ultimo_relatorio
    print(f"🤖 Bot conectado como {bot.user}")
    # Procurar o arquivo Excel mais recente
    arquivos = glob.glob("Relatorio_ *.xlsx")
    if arquivos:
        arquivo_ultimo_relatorio = max(arquivos, key=os.path.getmtime)
        print(f"[DEBUG] Último relatório Excel encontrado: {arquivo_ultimo_relatorio}")
    else:
        arquivo_ultimo_relatorio = None
        print("[DEBUG] Nenhum relatório Excel encontrado ao iniciar.")
    # Registrar views persistentes para todos os comunicados antigos
    for json_file in os.listdir("relatorios"):
        if json_file.endswith(".json") and json_file != "relatorio_atual.json":
            titulo = json_file[:-5]  # remove .json
            bot.add_view(ComunicadoViewPersistente(titulo))
            print(f"[DEBUG] View persistente registrada para comunicado: {titulo}")
    await bot.tree.sync()

CARGOS_PERMITIDOS = [
    "RH",
    "CEO",
    "Head Financeiro",
    "Planejamento",
    "Qualidade",
    "Supervisão"
]

@bot.tree.command(name="comunicado", description="Enviar um comunicado oficial")
@app_commands.describe(
    anexo1="Primeiro anexo (opcional)",
    anexo2="Segundo anexo (opcional)",
    anexo3="Terceiro anexo (opcional)",
    anexo4="Quarto anexo (opcional)",
    anexo5="Quinto anexo (opcional)"
)
async def comunicado(
    interaction: discord.Interaction,
    anexo1: Optional[discord.Attachment] = None,
    anexo2: Optional[discord.Attachment] = None,
    anexo3: Optional[discord.Attachment] = None,
    anexo4: Optional[discord.Attachment] = None,
    anexo5: Optional[discord.Attachment] = None
):
    # Verificação de cargo
    if not any(role.name in CARGOS_PERMITIDOS for role in getattr(interaction.user, 'roles', [])):
        await interaction.response.send_message("❌ Você não tem permissão para enviar comunicados.", ephemeral=True)
        return
    modal = NovoComunicadoModal()
    # Passa os anexos para o modal
    modal.anexos = [a for a in [anexo1, anexo2, anexo3, anexo4, anexo5] if a is not None]
    await interaction.response.send_modal(modal)

# Ao criar um novo comunicado, armazene o ID da mensagem
class NovoComunicadoModal(discord.ui.Modal, title="📝 Criar Comunicado"):
    def __init__(self):
        super().__init__()
        self.titulo = discord.ui.TextInput(label="Título", placeholder="Ex: Mudança de horário", required=True)
        self.conteudo = discord.ui.TextInput(label="Conteúdo", style=discord.TextStyle.long, placeholder="Escreva aqui o comunicado completo...", required=True)
        self.add_item(self.titulo)
        self.add_item(self.conteudo)
        self.anexos = []  # Lista de anexos, será preenchida pelo comando

    async def on_submit(self, interaction: discord.Interaction):
        global titulo_ultimo_comunicado, arquivo_ultimo_relatorio, ultima_msg_comunicado_id
        canal = interaction.client.get_channel(CANAL_ID)
        if canal:
            titulo = self.titulo.value
            conteudo = self.conteudo.value
            titulo_ultimo_comunicado = titulo
            nome_arquivo = f"Relatorio_ {titulo}.xlsx"
            arquivo_ultimo_relatorio = nome_arquivo
            print(f"[DEBUG] Criando arquivo Excel: {nome_arquivo}")
            guild = interaction.guild
            wb = Workbook()
            ws = wb.active
            ws.append(["Membro", "Status"])
            membros_status = []
            for member in guild.members:
                if not member.bot:
                    ws.append([member.display_name, "Não viu"])
                    membros_status.append({"id": member.id, "nome": member.display_name, "status": "Não viu"})
            wb.save(nome_arquivo)
            salvar_relatorio_json(titulo, membros_status, conteudo)
            view = ComunicadoViewPersistente(titulo)
            files = []
            if hasattr(self, 'anexos') and self.anexos:
                for anexo in self.anexos:
                    files.append(await anexo.to_file())
            msg = await canal.send(
                f"🔔 Novo comunicado: **{titulo}** \n"
                f"enviado por {interaction.user.display_name}",
                view=view,
                files=files if files else None
            )
            ultima_msg_comunicado_id = msg.id
            await interaction.response.send_message(f"✅ Comunicado publicado com sucesso! Relatório gerado: {nome_arquivo}", ephemeral=True)
        else:
            await interaction.response.send_message("❌ Canal não encontrado. Verifique o ID.", ephemeral=True)

# Comando para apagar o último comunicado enviado
@bot.tree.command(name="apagar_comunicado", description="Apaga o último comunicado enviado")
async def apagar_comunicado(interaction: discord.Interaction):
    # Verificação de cargo
    if not any(role.name in CARGOS_PERMITIDOS for role in getattr(interaction.user, 'roles', [])):
        await interaction.response.send_message("❌ Você não tem permissão para apagar comunicados.", ephemeral=True)
        return
    global ultima_msg_comunicado_id
    canal = interaction.client.get_channel(CANAL_ID)
    if not canal or not ultima_msg_comunicado_id:
        await interaction.response.send_message("❌ Nenhum comunicado encontrado para apagar.", ephemeral=True)
        return
    try:
        msg = await canal.fetch_message(ultima_msg_comunicado_id)
        await msg.delete()
        ultima_msg_comunicado_id = None
        await interaction.response.send_message("✅ Comunicado apagado com sucesso!", ephemeral=True)
    except Exception as e:
        await interaction.response.send_message(f"❌ Erro ao apagar comunicado: {e}", ephemeral=True)

# Remover o comando !relatorio
# Restaurar o comando de barra /relatorio

@bot.tree.command(name="relatorio", description="Enviar o relatório atualizado do último comunicado")
async def relatorio(interaction: discord.Interaction):
    print(f"[DEBUG] /relatorio chamado no canal {interaction.channel_id}")
    print(f"[DEBUG] CANAL_COMANDOS_ID: {CANAL_COMANDOS_ID}")
    global arquivo_ultimo_relatorio
    print(f"[DEBUG] arquivo_ultimo_relatorio: {arquivo_ultimo_relatorio}")
    if interaction.channel_id != CANAL_COMANDOS_ID:
        print("[DEBUG] Canal incorreto para o comando /relatorio.")
        await interaction.response.send_message("❌ Este comando só pode ser usado no canal de comandos.", ephemeral=True)
        return
    if not arquivo_ultimo_relatorio or not os.path.exists(arquivo_ultimo_relatorio):
        print(f"[DEBUG] Arquivo Excel não encontrado: {arquivo_ultimo_relatorio}")
        await interaction.response.send_message("❌ Nenhum relatório Excel encontrado.", ephemeral=True)
        return
    print(f"[DEBUG] Enviando arquivo: {arquivo_ultimo_relatorio}")
    await interaction.response.send_message("📄 Relatório atualizado (Excel):", ephemeral=False)
    await interaction.followup.send(file=discord.File(arquivo_ultimo_relatorio))

bot.run(TOKEN)
